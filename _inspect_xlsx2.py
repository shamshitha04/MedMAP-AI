import openpyxl
xlsx = r"c:\MedMAp-AI\dataset\interaction_counts_4.xlsx"
wb = openpyxl.load_workbook(xlsx)
for s in wb.sheetnames:
    ws = wb[s]
    print(f"Sheet: {s}, dims: {ws.dimensions}, max_row: {ws.max_row}, max_col: {ws.max_column}")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 15:
            break
        print("  ", row)
wb.close()
