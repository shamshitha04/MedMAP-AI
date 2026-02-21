import sqlite3, os, json

db = r"c:\MedMAp-AI\database\medmap.db"
print("=== medmap.db === Size:", os.path.getsize(db), "bytes")
c = sqlite3.connect(db)
cur = c.cursor()
tabs = cur.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
print("Tables:", [t[0] for t in tabs])
for t in tabs:
    n = t[0]
    cnt = cur.execute(f"SELECT COUNT(*) FROM [{n}]").fetchone()[0]
    cols = cur.execute(f"PRAGMA table_info([{n}])").fetchall()
    print(f"  {n}: {cnt} rows, cols: {[(c2[1], c2[2]) for c2 in cols]}")
    if cnt > 0 and cnt <= 50:
        rows = cur.execute(f"SELECT * FROM [{n}] LIMIT 5").fetchall()
        for r in rows:
            print("   ", r)
c.close()

# XLSX info
xlsx = r"c:\MedMAp-AI\dataset\interaction_counts_4.xlsx"
print(f"\n=== interaction_counts_4.xlsx === Size: {os.path.getsize(xlsx)} bytes")
try:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    for s in wb.sheetnames:
        ws = wb[s]
        print(f"  Sheet: {s}, dims: {ws.dimensions}")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 12:
                break
            print("   ", row)
    wb.close()
except ImportError:
    print("  openpyxl not installed - trying pandas")
    try:
        import pandas as pd
        df = pd.read_excel(xlsx, nrows=10)
        print(f"  Columns: {list(df.columns)}")
        print(f"  Shape: {df.shape}")
        print(df.head())
    except ImportError:
        print("  Neither openpyxl nor pandas available")
