import openpyxl
xlsx = r"c:\MedMAp-AI\dataset\interaction_counts_4.xlsx"
wb = openpyxl.load_workbook(xlsx, read_only=True)
for s in wb.sheetnames:
    ws = wb[s]
    print(f"Sheet: {s}, dims: {ws.dimensions}")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 15:
            break
        print("  ", row)
    # Count rows
    row_count = sum(1 for _ in ws.iter_rows())
    print(f"  Total rows: {row_count}")
wb.close()
